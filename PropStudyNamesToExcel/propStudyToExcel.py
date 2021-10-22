import re, os
from openpyxl import Workbook

# change working directory to the location of the textfile and desired output workbook
os.chdir(r"C:\Users\mechomicky\Burns & McDonnell\FirstEnergy FAN Team - General\RadioCompass Prop Study KMZ's")

# open and read the textfile with a list of prop studies
studiesFile = open('PropStudiesList.txt')
studiesText = studiesFile.read()

# create a regex to find OpCo, Site Name, Antenna Height, and Remote Height from each prop study name
mo = re.compile(r'(.*?)_(.*?)[\s_-]([0-9]{2,3})[\s_-]([0-9]{2})*')
matches = mo.findall(studiesText)

# create a regex to find antenna zone, if present
mo2 = re.compile(r'.*?_(.*?)([Bb][Ll][Uu][Ee]|[Rr][Ee][Dd]|[Ww][Hh][Ii][Tt][Ee]|[Oo][Mm][Nn][Ii])*[\s_-][0-9]{2,3}([Bb][Ll][Uu][Ee]|[Rr][Ee][Dd]|[Ww][Hh][Ii][Tt][Ee]|[Oo][Mm][Nn][Ii])*[\s_-](?:[0-9]{2})*[\s_-]*([Bb][Ll][Uu][Ee]|[Rr][Ee][Dd]|[Ww][Hh][Ii][Tt][Ee]|[Oo][Mm][Nn][Ii])*')
sectorMatches = mo2.findall(studiesText)

# modify the sectorMatches into a list so that blank strings can be removed
sectorList = []
for m in sectorMatches:
    sectorList.append(list(m))

for s in sectorList:
    i = 0
    while i < len(s):
        if s[i] == '':
            del s[i]
        else:
            i += 1

# create a new workbook object and set the active sheet
workbook = Workbook()
sheet = workbook.active

# create the column names
sheet.append(['OpCo/grouping', 'Site', 'Sector', 'Antenna Height', 'Remote Height'])

# add values from first matching to the sheet
for m in matches:
    sheet.append([m[0], m[1], '', m[2], m[3]])

# add values from second matching to sheet
for s in sectorList:
    if len(s) == 2:
        for row in range(2, sheet.max_row):
            if sheet.cell(row=row, column=2).value in s[0]:
                sheet.cell(row=row, column=3).value = s[1]

# save and close the workbook
workbook.save('Prop Study Height List.xlsx')
workbook.close